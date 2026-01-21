#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
أدوات تحليل الاستبيان وكشف الغش - نسخة Google Colab
ينشئ ملف Excel متقدم مع معادلات وتحليلات تلقائية
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

# Check if running in Colab
try:
    from google.colab import files
    IN_COLAB = True
    print("✅ تم اكتشاف بيئة Google Colab")
except ImportError:
    IN_COLAB = False
    print("ℹ️  ليس في Google Colab")


def upload_files_colab():
    """
    رفع الملفات في Google Colab
    """
    if not IN_COLAB:
        print("❌ هذه الوظيفة تعمل فقط في Google Colab")
        return None, None
    
    print("\n" + "="*80)
    print("📂 رفع الملفات في Google Colab")
    print("="*80)
    
    # Upload real data
    print("\n1️⃣ ارفع ملف البيانات الحقيقية:")
    print("   انقر على 'Choose Files' واختر ملف CSV أو Excel")
    uploaded_real = files.upload()
    
    if not uploaded_real:
        print("❌ لم يتم رفع ملف!")
        return None, None
    
    real_path = list(uploaded_real.keys())[0]
    print(f"   ✅ تم رفع: {real_path}")
    
    # Ask about fake data
    print("\n2️⃣ هل تريد رفع ملف بيانات مزيفة؟")
    print("   اكتب 'yes' لرفع ملف مزيف، أو اضغط Enter للتخطي")
    response = input("   👉 ").strip().lower()
    
    fake_path = None
    if response in ['yes', 'y', 'نعم']:
        print("\n   ارفع ملف البيانات المزيفة:")
        uploaded_fake = files.upload()
        if uploaded_fake:
            fake_path = list(uploaded_fake.keys())[0]
            print(f"   ✅ تم رفع: {fake_path}")
        else:
            print("   ⏭️  تم التخطي")
    else:
        print("   ⏭️  تم التخطي")
    
    return real_path, fake_path


def read_data_file(file_path):
    """
    قراءة ملف بيانات (CSV أو Excel)
    """
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.csv':
            # Try multiple encodings
            for encoding in ['utf-8-sig', 'utf-8', 'cp1256', 'windows-1256', 'latin1']:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    print(f"      ✓ {len(df)} صف (ترميز: {encoding})")
                    return df
                except:
                    continue
            raise ValueError("فشل قراءة CSV بجميع الترميزات")
        
        elif file_ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path)
            print(f"      ✓ {len(df)} صف")
            return df
        
        else:
            raise ValueError(f"نوع ملف غير مدعوم: {file_ext}")
    
    except Exception as e:
        print(f"      ✗ خطأ: {e}")
        return None


def analyze_data_quality(df):
    """
    تحليل جودة البيانات
    """
    likert_map = {
        'موافق بشدة': 5, 'موافق': 4, 'محايد': 3,
        'غير موافق': 2, 'غير موافق بشدة': 1
    }
    
    results = []
    
    for idx, row in df.iterrows():
        result = {
            'رقم الرد': idx + 1,
            'المصدر': row.get('مصدر البيانات', 'غير محدد')
        }
        
        try:
            # Find attention check columns
            q4_cols = [col for col in df.columns if 'محايد' in col and ('4' in col or 'الرابع' in col)]
            q7_cols = [col for col in df.columns if 'موافق بشدة' in col and ('7' in col or 'السابع' in col)]
            
            # Check Q4
            if q4_cols:
                result['نجح في السؤال 4'] = 'نعم' if row[q4_cols[0]] == 'محايد' else 'لا'
            else:
                result['نجح في السؤال 4'] = 'غ/م'
            
            # Check Q7
            if q7_cols:
                result['نجح في السؤال 7'] = 'نعم' if row[q7_cols[0]] == 'موافق بشدة' else 'لا'
            else:
                result['نجح في السؤال 7'] = 'غ/م'
            
            # Find contradiction columns
            frustration_cols = [col for col in df.columns if 'إحباط' in col]
            happiness_cols = [col for col in df.columns if 'سعادة' in col]
            waste_cols = [col for col in df.columns if 'هدر' in col or 'تهدر' in col]
            control_cols = [col for col in df.columns if 'تحكم' in col]
            
            # Contradiction 1
            if frustration_cols and happiness_cols:
                q6_val = likert_map.get(row[frustration_cols[0]], 0)
                q10_val = likert_map.get(row[happiness_cols[0]], 0)
                result['تناقض (إحباط+سعادة)'] = 'نعم' if (q6_val >= 4 and q10_val >= 4) else 'لا'
            else:
                result['تناقض (إحباط+سعادة)'] = 'غ/م'
            
            # Contradiction 2
            if waste_cols and control_cols:
                q8_val = likert_map.get(row[waste_cols[0]], 0)
                q9_val = likert_map.get(row[control_cols[0]], 0)
                result['تناقض (وقت+تحكم)'] = 'نعم' if (q8_val >= 4 and q9_val >= 4) else 'لا'
            else:
                result['تناقض (وقت+تحكم)'] = 'غ/م'
            
            # Standard deviation
            likert_cols = [col for col in df.columns if any(val in str(row[col]) for val in likert_map.keys())]
            
            if len(likert_cols) >= 5:
                responses = [likert_map.get(row[col], 0) for col in likert_cols if row[col] in likert_map]
                if len(responses) >= 5:
                    std = np.std(responses, ddof=1)
                    result['الانحراف المعياري'] = round(std, 2)
                    result['انحراف منخفض'] = 'نعم' if std < 0.5 else 'لا'
                else:
                    result['الانحراف المعياري'] = 'غ/ح'
                    result['انحراف منخفض'] = 'لا'
            else:
                result['الانحراف المعياري'] = 'غ/ح'
                result['انحراف منخفض'] = 'لا'
            
            # Final assessment
            issues = sum([
                result.get('نجح في السؤال 4') == 'لا',
                result.get('نجح في السؤال 7') == 'لا',
                result.get('تناقض (إحباط+سعادة)') == 'نعم',
                result.get('تناقض (وقت+تحكم)') == 'نعم',
                result.get('انحراف منخفض') == 'نعم'
            ])
            
            if issues == 0:
                result['التقييم النهائي'] = '✅ نظيف'
            elif issues <= 2:
                result['التقييم النهائي'] = '⚠️ مشبوه'
            else:
                result['التقييم النهائي'] = '❌ مزيف'
        
        except Exception as e:
            result['التقييم النهائي'] = '⚠️ خطأ'
        
        results.append(result)
    
    return pd.DataFrame(results)


def create_demographics_summary(df):
    """
    ملخص الإحصائيات
    """
    summary = []
    
    demographics_patterns = {
        'الجنس': ['جنس', 'Gender'],
        'العمر': ['عمر', 'Age'],
        'التعليم': ['تعليم', 'المستوى', 'Education'],
        'الوظيفة': ['وظيف', 'عمل', 'Job', 'Employment'],
        'ساعات الاستخدام': ['ساعات', 'استخدام', 'Usage', 'Hours']
    }
    
    for var_name, patterns in demographics_patterns.items():
        matching_cols = []
        for col in df.columns:
            if any(pattern in col for pattern in patterns):
                matching_cols.append(col)
        
        if matching_cols:
            col = matching_cols[0]
            counts = df[col].value_counts()
            for value, count in counts.items():
                summary.append({
                    'المتغير': var_name,
                    'القيمة': str(value),
                    'العدد': int(count),
                    'النسبة %': round(count / len(df) * 100, 1)
                })
    
    if not summary:
        summary.append({
            'المتغير': 'تنبيه',
            'القيمة': 'لم يتم العثور على بيانات ديموغرافية',
            'العدد': 0,
            'النسبة %': 0.0
        })
    
    return pd.DataFrame(summary)


def create_user_guide():
    """
    دليل الاستخدام
    """
    return pd.DataFrame([
        {'القسم': '📌 مقدمة', 'الشرح': 'ملف تحليل شامل لجودة بيانات الاستبيان'},
        {'القسم': '📊 البيانات الكاملة', 'الشرح': 'جميع البيانات (حقيقية + مزيفة إن وجدت)'},
        {'القسم': '✅ البيانات الحقيقية', 'الشرح': 'البيانات الأصلية من المشاركين'},
        {'القسم': '❌ البيانات المزيفة', 'الشرح': 'بيانات محاكاة لأغراض البحث (إن وجدت)'},
        {'القسم': '🔍 تحليل الجودة', 'الشرح': 'تقييم تفصيلي لكل رد: نظيف / مشبوه / مزيف'},
        {'القسم': '📈 الإحصائيات', 'الشرح': 'توزيع المتغيرات الديموغرافية'},
        {'القسم': '✓ نظيف', 'الشرح': 'رد عالي الجودة، اجتاز جميع الفحوصات'},
        {'القسم': '⚠ مشبوه', 'الشرح': 'رد يحتاج مراجعة، فشل في 1-2 فحوصات'},
        {'القسم': '✗ مزيف', 'الشرح': 'رد منخفض الجودة، فشل في 3+ فحوصات'},
        {'القسم': '💡 كيفية الاستخدام', 'الشرح': '1) افتح "تحليل الجودة" 2) راجع الردود المشبوهة 3) قرر الاستبعاد'},
        {'القسم': '⚠️ ملاحظة', 'الشرح': 'وضّح في البحث أن البيانات المزيفة للتوضيح فقط'}
    ])


def format_excel(filename):
    """
    تنسيق ملف Excel
    """
    wb = load_workbook(filename)
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    # Color code quality sheet
    if 'تحليل الجودة' in wb.sheetnames:
        ws = wb['تحليل الجودة']
        
        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            assessment = str(row[-1].value)
            
            if 'نظيف' in assessment:
                for cell in row:
                    cell.fill = green
            elif 'مشبوه' in assessment:
                for cell in row:
                    cell.fill = yellow
            elif 'مزيف' in assessment:
                for cell in row:
                    cell.fill = red
    
    wb.save(filename)


def main():
    print("\n" + "="*80)
    print("📊 تحليل جودة بيانات الاستبيان - Survey Quality Analyzer")
    print("="*80)
    if IN_COLAB:
        print("النسخة 4.0 - Google Colab Edition")
    else:
        print("النسخة 4.0 - Standalone Edition")
    print("="*80)
    
    # Upload or select files
    if IN_COLAB:
        real_path, fake_path = upload_files_colab()
    else:
        print("\n⚠️  هذا الكود مصمم لـ Google Colab")
        print("للاستخدام على جهازك، استخدم النسخة GUI بدلاً من ذلك")
        return
    
    if real_path is None:
        print("\n❌ تم الإلغاء")
        return
    
    # Read files
    print("\n📖 قراءة الملفات...")
    print("="*80)
    
    print("\n  [1] البيانات الحقيقية:")
    df_real = read_data_file(real_path)
    
    if df_real is None:
        print("\n❌ فشل قراءة الملف")
        return
    
    df_real['مصدر البيانات'] = 'حقيقي'
    
    # Read fake data if provided
    df_fake = None
    if fake_path:
        print("\n  [2] البيانات المزيفة:")
        df_fake = read_data_file(fake_path)
        if df_fake is not None:
            df_fake['مصدر البيانات'] = 'مزيف (محاكاة)'
    
    # Combine data
    if df_fake is not None:
        print("\n🔀 دمج البيانات...")
        df_combined = pd.concat([df_real, df_fake], ignore_index=True)
        print(f"   ✓ المجموع: {len(df_combined)} رد")
    else:
        df_combined = df_real.copy()
    
    # Analyze
    print("\n🔍 تحليل البيانات...")
    print("="*80)
    
    print("\n  ⚡ تحليل الجودة...")
    quality_df = analyze_data_quality(df_combined)
    
    print("  ⚡ الإحصائيات الديموغرافية...")
    demographics_df = create_demographics_summary(df_combined)
    
    print("  ⚡ دليل الاستخدام...")
    guide_df = create_user_guide()
    
    # Create Excel
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"تحليل_الاستبيان_{timestamp}.xlsx"
    
    print("\n📝 إنشاء ملف Excel...")
    print("="*80)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            print("\n  📄 البيانات الكاملة...")
            df_combined.to_excel(writer, sheet_name='البيانات الكاملة', index=False)
            
            print("  📄 البيانات الحقيقية...")
            df_real.to_excel(writer, sheet_name='البيانات الحقيقية', index=False)
            
            if df_fake is not None:
                print("  📄 البيانات المزيفة...")
                df_fake.to_excel(writer, sheet_name='البيانات المزيفة', index=False)
            
            print("  📄 تحليل الجودة...")
            quality_df.to_excel(writer, sheet_name='تحليل الجودة', index=False)
            
            print("  📄 الإحصائيات...")
            demographics_df.to_excel(writer, sheet_name='الإحصائيات', index=False)
            
            print("  📄 دليل الاستخدام...")
            guide_df.to_excel(writer, sheet_name='دليل الاستخدام', index=False)
        
        print("\n🎨 تنسيق الملف...")
        format_excel(output_path)
        
        print("\n" + "="*80)
        print("✅ تم بنجاح!")
        print("="*80)
        print(f"\n📁 اسم الملف: {output_path}")
        
        # Summary stats
        print("\n📊 ملخص النتائج:")
        print("-"*80)
        clean = len(quality_df[quality_df['التقييم النهائي'].str.contains('نظيف', na=False)])
        suspicious = len(quality_df[quality_df['التقييم النهائي'].str.contains('مشبوه', na=False)])
        fake = len(quality_df[quality_df['التقييم النهائي'].str.contains('مزيف', na=False)])
        
        print(f"  ✅ نظيفة: {clean} ({clean/len(quality_df)*100:.1f}%)")
        print(f"  ⚠️  مشبوهة: {suspicious} ({suspicious/len(quality_df)*100:.1f}%)")
        print(f"  ❌ مزيفة: {fake} ({fake/len(quality_df)*100:.1f}%)")
        
        # Download file in Colab
        if IN_COLAB:
            print("\n📥 تحميل الملف...")
            files.download(output_path)
            print("✅ تم! يمكنك العثور على الملف في مجلد Downloads")
    
    except Exception as e:
        print(f"\n❌ خطأ: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "="*80)
    print("انتهى التحليل!")
    print("="*80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ خطأ غير متوقع: {e}")
        import traceback
        traceback.print_exc()
