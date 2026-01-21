#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
أدوات تحليل الاستبيان وكشف الغش - مع واجهة اختيار الملفات
ينشئ ملف Excel متقدم مع معادلات وتحليلات تلقائية
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

# Import tkinter
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
    print("⚠️  تحذير: مكتبة tkinter غير متوفرة")


def select_files():
    """
    اختيار الملفات عبر نافذة مرئية
    """
    if not TKINTER_AVAILABLE:
        print("❌ الواجهة الرسومية غير متوفرة")
        return None, None, None
    
    root = tk.Tk()
    root.withdraw()
    
    try:
        root.attributes('-topmost', True)
    except:
        pass
    
    print("\n" + "="*80)
    print("📂 اختيار الملفات")
    print("="*80)
    
    # Step 1: Real data
    print("\n1️⃣ اختر ملف البيانات الحقيقية...")
    messagebox.showinfo(
        "خطوة 1 من 3",
        "اختر ملف البيانات الحقيقية\n\n"
        "الملف الذي يحتوي على ردود المشاركين الفعلية"
    )
    
    real_path = filedialog.askopenfilename(
        title="اختر ملف البيانات الحقيقية",
        filetypes=[
            ("Excel/CSV files", "*.xlsx *.xls *.csv"),
            ("Excel files", "*.xlsx *.xls"),
            ("CSV files", "*.csv"),
            ("All files", "*.*")
        ]
    )
    
    if not real_path:
        messagebox.showwarning("تحذير", "لم يتم اختيار ملف!\n\nالبرنامج سينتهي.")
        root.destroy()
        return None, None, None
    
    print(f"   ✅ {os.path.basename(real_path)}")
    
    # Step 2: Fake data (optional)
    print("\n2️⃣ ملف البيانات المزيفة (اختياري)...")
    
    response = messagebox.askyesno(
        "خطوة 2 من 3",
        "هل لديك ملف بيانات مزيفة للمقارنة؟\n\n"
        "البيانات المزيفة تُستخدم لأغراض البحث والتوضيح\n\n"
        "• Yes = لاختيار ملف مزيف\n"
        "• No = المتابعة بدون ملف مزيف"
    )
    
    fake_path = None
    if response:
        fake_path = filedialog.askopenfilename(
            title="اختر ملف البيانات المزيفة (اختياري)",
            filetypes=[
                ("Excel/CSV files", "*.xlsx *.xls *.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if fake_path:
            print(f"   ✅ {os.path.basename(fake_path)}")
        else:
            print("   ⏭️  تم التخطي")
    else:
        print("   ⏭️  تم التخطي")
    
    # Step 3: Output location
    print("\n3️⃣ اختر مكان حفظ ملف التحليل...")
    messagebox.showinfo(
        "خطوة 3 من 3",
        "اختر مكان واسم لحفظ ملف التحليل\n\n"
        "سيحتوي على:\n"
        "• البيانات الكاملة\n"
        "• تحليل الجودة\n"
        "• الإحصائيات\n"
        "• دليل الاستخدام"
    )
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"تحليل_الاستبيان_{timestamp}.xlsx"
    
    initial_dir = os.path.dirname(real_path) if real_path else os.getcwd()
    
    output_path = filedialog.asksaveasfilename(
        title="احفظ ملف التحليل",
        defaultextension=".xlsx",
        initialfile=default_name,
        initialdir=initial_dir,
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )
    
    if not output_path:
        output_path = os.path.join(initial_dir, default_name)
        print(f"   ⚠️  استخدام الاسم الافتراضي")
    
    print(f"   ✅ {os.path.basename(output_path)}")
    
    root.destroy()
    return real_path, fake_path, output_path


def read_data_file(file_path):
    """
    قراءة ملف بيانات (CSV أو Excel)
    """
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.csv':
            # Try multiple encodings
            for encoding in ['utf-8-sig', 'utf-8', 'cp1256', 'windows-1256', 'latin1']:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    print(f"      ✓ {len(df)} صف (ترميز: {encoding})")
                    return df
                except:
                    continue
            raise ValueError("فشل قراءة CSV بجميع الترميزات")
        
        elif file_ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path)
            print(f"      ✓ {len(df)} صف")
            return df
        
        else:
            raise ValueError(f"نوع ملف غير مدعوم: {file_ext}")
    
    except Exception as e:
        print(f"      ✗ خطأ: {e}")
        return None


def analyze_data_quality(df):
    """
    تحليل جودة البيانات
    """
    likert_map = {
        'موافق بشدة': 5, 'موافق': 4, 'محايد': 3,
        'غير موافق': 2, 'غير موافق بشدة': 1
    }
    
    results = []
    
    for idx, row in df.iterrows():
        result = {
            'رقم الرد': idx + 1,
            'المصدر': row.get('مصدر البيانات', 'غير محدد')
        }
        
        try:
            # Find attention check columns
            q4_cols = [col for col in df.columns if 'محايد' in col and ('4' in col or 'الرابع' in col)]
            q7_cols = [col for col in df.columns if 'موافق بشدة' in col and ('7' in col or 'السابع' in col)]
            
            # Check Q4
            if q4_cols:
                result['نجح في السؤال 4'] = 'نعم' if row[q4_cols[0]] == 'محايد' else 'لا'
            else:
                result['نجح في السؤال 4'] = 'غ/م'
            
            # Check Q7
            if q7_cols:
                result['نجح في السؤال 7'] = 'نعم' if row[q7_cols[0]] == 'موافق بشدة' else 'لا'
            else:
                result['نجح في السؤال 7'] = 'غ/م'
            
            # Find contradiction columns
            frustration_cols = [col for col in df.columns if 'إحباط' in col]
            happiness_cols = [col for col in df.columns if 'سعادة' in col]
            waste_cols = [col for col in df.columns if 'هدر' in col or 'تهدر' in col]
            control_cols = [col for col in df.columns if 'تحكم' in col]
            
            # Contradiction 1
            if frustration_cols and happiness_cols:
                q6_val = likert_map.get(row[frustration_cols[0]], 0)
                q10_val = likert_map.get(row[happiness_cols[0]], 0)
                result['تناقض (إحباط+سعادة)'] = 'نعم' if (q6_val >= 4 and q10_val >= 4) else 'لا'
            else:
                result['تناقض (إحباط+سعادة)'] = 'غ/م'
            
            # Contradiction 2
            if waste_cols and control_cols:
                q8_val = likert_map.get(row[waste_cols[0]], 0)
                q9_val = likert_map.get(row[control_cols[0]], 0)
                result['تناقض (وقت+تحكم)'] = 'نعم' if (q8_val >= 4 and q9_val >= 4) else 'لا'
            else:
                result['تناقض (وقت+تحكم)'] = 'غ/م'
            
            # Standard deviation
            likert_cols = [col for col in df.columns if any(val in str(row[col]) for val in likert_map.keys())]
            
            if len(likert_cols) >= 5:
                responses = [likert_map.get(row[col], 0) for col in likert_cols if row[col] in likert_map]
                if len(responses) >= 5:
                    std = np.std(responses, ddof=1)
                    result['الانحراف المعياري'] = round(std, 2)
                    result['انحراف منخفض'] = 'نعم' if std < 0.5 else 'لا'
                else:
                    result['الانحراف المعياري'] = 'غ/ح'
                    result['انحراف منخفض'] = 'لا'
            else:
                result['الانحراف المعياري'] = 'غ/ح'
                result['انحراف منخفض'] = 'لا'
            
            # Final assessment
            issues = sum([
                result.get('نجح في السؤال 4') == 'لا',
                result.get('نجح في السؤال 7') == 'لا',
                result.get('تناقض (إحباط+سعادة)') == 'نعم',
                result.get('تناقض (وقت+تحكم)') == 'نعم',
                result.get('انحراف منخفض') == 'نعم'
            ])
            
            if issues == 0:
                result['التقييم النهائي'] = '✅ نظيف'
            elif issues <= 2:
                result['التقييم النهائي'] = '⚠️ مشبوه'
            else:
                result['التقييم النهائي'] = '❌ مزيف'
        
        except Exception as e:
            result['التقييم النهائي'] = '⚠️ خطأ'
        
        results.append(result)
    
    return pd.DataFrame(results)


def create_demographics_summary(df):
    """
    ملخص الإحصائيات
    """
    summary = []
    
    demographics_patterns = {
        'الجنس': ['جنس', 'Gender'],
        'العمر': ['عمر', 'Age'],
        'التعليم': ['تعليم', 'المستوى', 'Education'],
        'الوظيفة': ['وظيف', 'عمل', 'Job', 'Employment'],
        'ساعات الاستخدام': ['ساعات', 'استخدام', 'Usage', 'Hours']
    }
    
    for var_name, patterns in demographics_patterns.items():
        matching_cols = []
        for col in df.columns:
            if any(pattern in col for pattern in patterns):
                matching_cols.append(col)
        
        if matching_cols:
            col = matching_cols[0]
            counts = df[col].value_counts()
            for value, count in counts.items():
                summary.append({
                    'المتغير': var_name,
                    'القيمة': str(value),
                    'العدد': int(count),
                    'النسبة %': round(count / len(df) * 100, 1)
                })
    
    if not summary:
        summary.append({
            'المتغير': 'تنبيه',
            'القيمة': 'لم يتم العثور على بيانات ديموغرافية',
            'العدد': 0,
            'النسبة %': 0.0
        })
    
    return pd.DataFrame(summary)


def create_user_guide():
    """
    دليل الاستخدام
    """
    return pd.DataFrame([
        {'القسم': '📌 مقدمة', 'الشرح': 'ملف تحليل شامل لجودة بيانات الاستبيان'},
        {'القسم': '📊 البيانات الكاملة', 'الشرح': 'جميع البيانات (حقيقية + مزيفة إن وجدت)'},
        {'القسم': '✅ البيانات الحقيقية', 'الشرح': 'البيانات الأصلية من المشاركين'},
        {'القسم': '❌ البيانات المزيفة', 'الشرح': 'بيانات محاكاة لأغراض البحث (إن وجدت)'},
        {'القسم': '🔍 تحليل الجودة', 'الشرح': 'تقييم تفصيلي لكل رد: نظيف / مشبوه / مزيف'},
        {'القسم': '📈 الإحصائيات', 'الشرح': 'توزيع المتغيرات الديموغرافية'},
        {'القسم': '✓ نظيف', 'الشرح': 'رد عالي الجودة، اجتاز جميع الفحوصات'},
        {'القسم': '⚠ مشبوه', 'الشرح': 'رد يحتاج مراجعة، فشل في 1-2 فحوصات'},
        {'القسم': '✗ مزيف', 'الشرح': 'رد منخفض الجودة، فشل في 3+ فحوصات'},
        {'القسم': '💡 كيفية الاستخدام', 'الشرح': '1) افتح "تحليل الجودة" 2) راجع الردود المشبوهة 3) قرر الاستبعاد'},
        {'القسم': '⚠️ ملاحظة', 'الشرح': 'وضّح في البحث أن البيانات المزيفة للتوضيح فقط'}
    ])


def format_excel(filename):
    """
    تنسيق ملف Excel
    """
    wb = load_workbook(filename)
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    # Color code quality sheet
    if 'تحليل الجودة' in wb.sheetnames:
        ws = wb['تحليل الجودة']
        
        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            assessment = str(row[-1].value)
            
            if 'نظيف' in assessment:
                for cell in row:
                    cell.fill = green
            elif 'مشبوه' in assessment:
                for cell in row:
                    cell.fill = yellow
            elif 'مزيف' in assessment:
                for cell in row:
                    cell.fill = red
    
    wb.save(filename)


def main():
    print("\n" + "="*80)
    print("📊 تحليل جودة بيانات الاستبيان - Survey Quality Analyzer")
    print("="*80)
    print("النسخة 3.0 - مع واجهة اختيار الملفات")
    print("="*80)
    
    # Select files
    result = select_files()
    
    if result[0] is None:
        print("\n❌ تم الإلغاء")
        if TKINTER_AVAILABLE:
            input("\nاضغط Enter للخروج...")
        return
    
    real_path, fake_path, output_path = result
    
    # Read files
    print("\n📖 قراءة الملفات...")
    print("="*80)
    
    print("\n  [1] البيانات الحقيقية:")
    df_real = read_data_file(real_path)
    
    if df_real is None:
        if TKINTER_AVAILABLE:
            messagebox.showerror("خطأ", "فشل قراءة ملف البيانات الحقيقية!")
        print("\n❌ فشل")
        input("\nاضغط Enter للخروج...")
        return
    
    df_real['مصدر البيانات'] = 'حقيقي'
    
    # Read fake data if provided
    df_fake = None
    if fake_path:
        print("\n  [2] البيانات المزيفة:")
        df_fake = read_data_file(fake_path)
        if df_fake is not None:
            df_fake['مصدر البيانات'] = 'مزيف (محاكاة)'
    
    # Combine data
    if df_fake is not None:
        print("\n🔀 دمج البيانات...")
        df_combined = pd.concat([df_real, df_fake], ignore_index=True)
        print(f"   ✓ المجموع: {len(df_combined)} رد")
    else:
        df_combined = df_real.copy()
    
    # Analyze
    print("\n🔍 تحليل البيانات...")
    print("="*80)
    
    print("\n  ⚡ تحليل الجودة...")
    quality_df = analyze_data_quality(df_combined)
    
    print("  ⚡ الإحصائيات الديموغرافية...")
    demographics_df = create_demographics_summary(df_combined)
    
    print("  ⚡ دليل الاستخدام...")
    guide_df = create_user_guide()
    
    # Create Excel
    print("\n📝 إنشاء ملف Excel...")
    print("="*80)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            print("\n  📄 البيانات الكاملة...")
            df_combined.to_excel(writer, sheet_name='البيانات الكاملة', index=False)
            
            print("  📄 البيانات الحقيقية...")
            df_real.to_excel(writer, sheet_name='البيانات الحقيقية', index=False)
            
            if df_fake is not None:
                print("  📄 البيانات المزيفة...")
                df_fake.to_excel(writer, sheet_name='البيانات المزيفة', index=False)
            
            print("  📄 تحليل الجودة...")
            quality_df.to_excel(writer, sheet_name='تحليل الجودة', index=False)
            
            print("  📄 الإحصائيات...")
            demographics_df.to_excel(writer, sheet_name='الإحصائيات', index=False)
            
            print("  📄 دليل الاستخدام...")
            guide_df.to_excel(writer, sheet_name='دليل الاستخدام', index=False)
        
        print("\n🎨 تنسيق الملف...")
        format_excel(output_path)
        
        print("\n" + "="*80)
        print("✅ تم بنجاح!")
        print("="*80)
        print(f"\n📁 الموقع:\n   {output_path}")
        
        # Summary stats
        print("\n📊 ملخص النتائج:")
        print("-"*80)
        clean = len(quality_df[quality_df['التقييم النهائي'].str.contains('نظيف', na=False)])
        suspicious = len(quality_df[quality_df['التقييم النهائي'].str.contains('مشبوه', na=False)])
        fake = len(quality_df[quality_df['التقييم النهائي'].str.contains('مزيف', na=False)])
        
        print(f"  ✅ نظيفة: {clean} ({clean/len(quality_df)*100:.1f}%)")
        print(f"  ⚠️  مشبوهة: {suspicious} ({suspicious/len(quality_df)*100:.1f}%)")
        print(f"  ❌ مزيفة: {fake} ({fake/len(quality_df)*100:.1f}%)")
        
        if TKINTER_AVAILABLE:
            messagebox.showinfo(
                "نجاح ✅",
                f"تم إنشاء ملف التحليل بنجاح!\n\n"
                f"📊 الإحصائيات:\n"
                f"• نظيفة: {clean} ({clean/len(quality_df)*100:.1f}%)\n"
                f"• مشبوهة: {suspicious} ({suspicious/len(quality_df)*100:.1f}%)\n"
                f"• مزيفة: {fake} ({fake/len(quality_df)*100:.1f}%)\n\n"
                f"📁 الموقع:\n{os.path.dirname(output_path)}"
            )
    
    except Exception as e:
        print(f"\n❌ خطأ: {e}")
        if TKINTER_AVAILABLE:
            messagebox.showerror("خطأ", f"فشل إنشاء الملف:\n{e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "="*80)
    input("\nاضغط Enter للخروج...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ خطأ غير متوقع: {e}")
        import traceback
        traceback.print_exc()
        input("\nاضغط Enter للخروج...")
